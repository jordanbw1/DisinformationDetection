import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import math


def compute_sheet_stats(inFileName, outFileName):
    # Define file paths
    script_directory = os.path.dirname(os.path.abspath(__file__))
    parent_directory = os.path.dirname(script_directory)
    inFileLoc = os.path.join(parent_directory, "dynamic", "prompt_results", inFileName)
    outFileLoc = os.path.join(parent_directory, "dynamic", "prompt_results", outFileName)

    # Read data from CSV into DataFrame
    df = pd.read_csv(inFileLoc, header=0)

    # Creating the pivot table
    pivot_table = pd.pivot_table(df, values='id', index='label', columns='response', aggfunc='count', fill_value=0)

    # Adding a total column and row
    pivot_table['Total'] = pivot_table.sum(axis=1)
    pivot_table.loc['Total'] = pivot_table.sum()

    # Get values
    try:
        correctDisinfo = pivot_table.at[1, 1]
    except KeyError:
        correctDisinfo = 0
    try:
        correctNoDisinfo = pivot_table.at[0, 0]
    except KeyError:
        correctNoDisinfo = 0
    try:
        falseNegative = pivot_table.at[1, 0]
    except KeyError:
        falseNegative = 0
    try:
        falsePositive = pivot_table.at[0, 1]
    except KeyError:
        falsePositive = 0

    # Calculate statistics
    # accuracy = (correctDisinfo + correctNoDisinfo) / pivot_table.loc['Total', 'Total']
    # precision = correctDisinfo / (correctDisinfo + falsePositive)
    # recall = correctDisinfo / (correctDisinfo + falseNegative)
    # f1_score = (2 * precision * recall) / (precision + recall)
    # tpr = correctDisinfo / (correctDisinfo + falseNegative) # Compute True Positive Rate (TPR)
    # fpr = falsePositive / (falsePositive + correctNoDisinfo) # Compute False Positive Rate (FPR)
    # fnr = falseNegative / (falseNegative + correctDisinfo) # Compute False Negative Rate (FNR)
    # tnr = correctNoDisinfo / (correctNoDisinfo + falsePositive) # Compute True Negative Rate (TNR)
    # Compute Accuracy
    accuracy_denominator = pivot_table.loc['Total', 'Total']
    accuracy = (correctDisinfo + correctNoDisinfo) / accuracy_denominator if accuracy_denominator != 0 else 0

    # Compute Precision
    precision_denominator = correctDisinfo + falsePositive
    precision = correctDisinfo / precision_denominator if precision_denominator != 0 else 0

    # Compute Recall
    recall_denominator = correctDisinfo + falseNegative
    recall = correctDisinfo / recall_denominator if recall_denominator != 0 else 0

    # Compute F1 Score
    f1_score_denominator = precision + recall
    f1_score = (2 * precision * recall) / f1_score_denominator if f1_score_denominator != 0 else 0

    # Compute True Positive Rate (TPR)
    tpr_denominator = correctDisinfo + falseNegative
    tpr = correctDisinfo / tpr_denominator if tpr_denominator != 0 else 0

    # Compute False Positive Rate (FPR)
    fpr_denominator = falsePositive + correctNoDisinfo
    fpr = falsePositive / fpr_denominator if fpr_denominator != 0 else 0

    # Compute False Negative Rate (FNR)
    fnr_denominator = falseNegative + correctDisinfo
    fnr = falseNegative / fnr_denominator if fnr_denominator != 0 else 0

    # Compute True Negative Rate (TNR)
    tnr_denominator = correctNoDisinfo + falsePositive
    tnr = correctNoDisinfo / tnr_denominator if tnr_denominator != 0 else 0

    # Round numbers
    accuracy = round(accuracy, 3)
    precision = round(precision, 3)
    recall = round(recall, 3)
    f1_score = round(f1_score, 3)
    tpr = round(tpr, 3)
    fpr = round(fpr, 3)
    fnr = round(fnr, 3)
    tnr = round(tnr, 3)

    # Pivot table for confidence_level
    pivot_table_confidence = pd.pivot_table(df, values='id', index='confidence_level', aggfunc='count', fill_value=0)

    # Pivot table for truth_level
    pivot_table_truth = pd.pivot_table(df, values='id', index='truth_level', aggfunc='count', fill_value=0)

    # Create ExcelWriter object
    with pd.ExcelWriter(outFileLoc, engine='openpyxl') as writer:
        # Write DataFrame to the first sheet named "Results"
        df.to_excel(writer, sheet_name='Results', index=False)

    # Load the workbook after it has been saved
    wb = load_workbook(outFileLoc)

    # Create a new sheet named "Pivots"
    pivot_sheet = wb.create_sheet(title='Stats')

    # Set the width of columns manualy
    pivot_sheet.column_dimensions['A'].width = 40
    pivot_sheet.column_dimensions['B'].width = 20
    pivot_sheet.column_dimensions['C'].width = 20
    pivot_sheet.column_dimensions['D'].width = 20

    # Write individual values to the "Pivots" sheet
    values = {
        'Correct Disinformation': correctDisinfo,
        'Correct No Disinformation': correctNoDisinfo,
        'False Negative (missed disinformation)': falseNegative,
        'False Positive (claimed disinfo when it wasn\'t)': falsePositive,
        'Accuracy': accuracy,
        'Precision': precision,
        'Recall': recall,
        'F1_score': f1_score,
        'True Positive Rate (TPR)': tpr,
        'False Positive Rate (FPR)': fpr,
        'False Negative Rate (FNR)': fnr,
        'True Negative Rate (TNR)': tnr
    }
    print("VALS", values)

    # Add a starting explanation row
    pivot_sheet.cell(row=1, column=1, value="Summary statistics")
    pivot_header_row_indices = [pivot_sheet.max_row]

    for i, (key, value) in enumerate(values.items()):
        pivot_sheet.cell(row=i+2, column=1, value=key)
        pivot_sheet.cell(row=i+2, column=2, value=value)

    # Add border for end of summary
    pivot_header_row_indices.append(pivot_sheet.max_row+1)

    # Add empty rows
    pivot_sheet.append([])
    pivot_sheet.append([])
    pivot_sheet.append([])
    pivot_sheet.append([])

    # Prepare Pivot headers
    pivot_sheet.append(["Confusion Matrix for Accuracy"])
    pivot_header_row_indices.append(pivot_sheet.max_row)
    pivot_sheet.append(["Dataset Truth Values", "AI Claimed False", "AI Claimed True", "Total"])
    pivot_header_row_indices.append(pivot_sheet.max_row)
    # Write pivot tables for label, confidence_level, and truth_level beneath the individual values
    for r in pd.DataFrame(pivot_table).iterrows():
        row_label = "Grand Total"
        if str(r[0]) == "0":
            row_label = "Actually False"
        if str(r[0]) == "1":
            row_label = "Actually True"
        # Append row
        pivot_sheet.append([row_label]+r[1].tolist())
    pivot_header_row_indices.append(pivot_sheet.max_row)

    # Add empty rows
    pivot_sheet.append([])
    pivot_sheet.append([])
    pivot_sheet.append([])

    # Pivot for confidence level
    # Prepare Headers
    pivot_sheet.append(["Confusion Matrix for Confidence Level"])
    pivot_header_row_indices.append(pivot_sheet.max_row)
    pivot_sheet.append(["AI Confidence Rating", "Count of Responses"])
    pivot_header_row_indices.append(pivot_sheet.max_row)
    # Append Rows
    confidence_total = 0
    for r in pd.DataFrame(pivot_table_confidence).iterrows():
        pivot_sheet.append([r[0]]+r[1].tolist())
        confidence_total += sum(value for value in r[1].tolist())
    # Add Grand Total row
    pivot_sheet.append(["Grand Total", confidence_total])
    pivot_header_row_indices.append(pivot_sheet.max_row)

    # Add empty rows
    pivot_sheet.append([])
    pivot_sheet.append([])
    pivot_sheet.append([])

    pivot_sheet.append(["Pivot Table for truth_level"])
    pivot_header_row_indices.append(pivot_sheet.max_row)
    pivot_sheet.append(["AI Confidence Rating", "Count of Responses"])
    pivot_header_row_indices.append(pivot_sheet.max_row)
    # Append Rows
    truth_total = 0
    for r in pd.DataFrame(pivot_table_truth).iterrows():
        pivot_sheet.append([r[0]]+r[1].tolist())
        truth_total += sum(value for value in r[1].tolist())
    # Add Grand Total row
    pivot_sheet.append(["Grand Total", truth_total])
    pivot_header_row_indices.append(pivot_sheet.max_row)

    # Apply fill color only to pivot header rows
    for row_index in pivot_header_row_indices:
        for cell in pivot_sheet[row_index]:
            cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
            cell.font = Font(bold=True)

    # Save the workbook
    wb.save(outFileLoc)

    # Make stats variable
    numCorrect = correctDisinfo + correctNoDisinfo
    numRows = df.shape[0]

    stats = {
        'tPos': correctDisinfo,
        'tNeg': correctNoDisinfo,
        'fNeg': falseNegative,
        'fPos': falsePositive,
        'accuracy': round(accuracy, 2),
        'precision': round(precision, 2),
        'recall': round(recall, 2),
        'fscore': round(f1_score, 2),
        'TPR': tpr,
        'FPR': fpr,
        'FNR': fnr,
        'TNR': tnr,
        'num_rows': numRows,
        'num_correct': numCorrect,
        'percent_correct': math.floor((numCorrect / numRows) * 100),
        'percent_TPR': round((correctDisinfo / numRows) * 100),
        'percent_FPR': round((falsePositive / numRows) * 100),
        'percent_TNR': round((correctNoDisinfo / numRows) * 100),
        'percent_FNR': round((falseNegative / numRows) * 100)
    }
    print(stats)

    # Return the filename and stats
    return stats
